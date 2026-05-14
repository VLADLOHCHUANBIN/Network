import re
import os
import shutil
from openpyxl import load_workbook

# ==========================================
# 🛑 CONFIGURATION ZONE 🛑
# ==========================================
TEMPLATE_IP_KEYWORD = "172.28.XXX.XXX"
TEMPLATE_FILE_KEYWORD = "Security Checklist"
LOG_SUBFOLDER = "SCC"

# 🎯 YOUR EXACT COMMAND LIST (Cleaned for Huawei VRP)
COMMAND_LIST = [
    "dis cu | inc telnet",
    "dis ssh server status",
    "dis startup",
    "dis ntp sessions",
    "dis users",
    "dis cu | inc local-user",
    "dis info-center"
]

# 3. 🌐 MULTI-SITE ROUTER LIST & IP SLOT 🌐
MULTI_SITE_ROUTERS = {
    "A": {
        "B01": "aaa.aaa.aaa.aaa",
    },
    "B": {
        "B01": "bbb.bbb.bbb.bbb",
    },
    "C": {
        "C01": "ccc.ccc.ccc.ccc",
    }
}


# ==========================================

# --- Precision Log Slicer (Regex Version) ---
def extract_specific_command(log_text, command, router_name):
    safe_command = re.escape(command)
    pattern = rf"{safe_command}[ \t\r]*\n?(.*?)(?=\n<|\n\[|\Z)"
    match = re.search(pattern, log_text, flags=re.DOTALL)

    if match:
        raw_output = match.group(1).strip()
        return f"<{router_name}> {command}\n{raw_output}"
    else:
        return f"[Command '{command}' not found in log]"


# --- Smart Excel Update Function ---
def update_excel_document(filepath, router_name, ip_address, log_text):
    try:
        wb = load_workbook(filepath)
        changes_made = False

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):

                        # 1. Update the IP address if it finds the placeholder
                        if TEMPLATE_IP_KEYWORD in cell.value:
                            cell.value = cell.value.replace(TEMPLATE_IP_KEYWORD, ip_address)
                            changes_made = True

                        # 2. Check if this cell contains one of your exact commands
                        for command in COMMAND_LIST:
                            if command.strip().lower() in cell.value.strip().lower():
                                specific_output = extract_specific_command(log_text, command, router_name)

                                # ✨ Overwrite the exact cell with the new output ✨
                                cell.value = specific_output
                                changes_made = True

        if changes_made:
            wb.save(filepath)
            return True
        return False
    except Exception as e:
        print(f"    [!] Error updating Excel doc: {e}")
        return False


# --- Safe File Reader ---
def safe_read_log(filepath):
    encodings = ['utf-8', 'utf-16', 'windows-1252', 'latin-1']
    for enc in encodings:
        try:
            with open(filepath, 'r', encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
        return f.read()


# --- Main SCC Automation ---
def process_scc_excel():
    base_dir = os.getcwd()
    template_dir_path = os.path.join(base_dir, "xxx")
    outputs_base_folder = os.path.join(base_dir, "Command Outputs")

    if not os.path.exists(template_dir_path):
        print("Error: Could not find the 'xxx' template folder.")
        print("Please make sure your template folder is in the exact same location as this script.")
        return

    # 🌐 Loop through every site!
    for site_name, routers_data in MULTI_SITE_ROUTERS.items():
        site_folder_path = os.path.join(base_dir, site_name)
        if not os.path.exists(site_folder_path):
            os.makedirs(site_folder_path)

        print(f"\n==================================================")
        print(f"🚀 Starting Final SCC Deployment for {site_name}")
        print(f"==================================================")

        for router_name, ip_address in routers_data.items():
            new_folder_path = os.path.join(site_folder_path, router_name)

            if not os.path.exists(new_folder_path):
                shutil.copytree(template_dir_path, new_folder_path)
                print(f"\n[+] Created folder and copying template for: {router_name}")

                for filename in os.listdir(new_folder_path):
                    if filename.startswith("~$"):
                        continue

                    # Rename the file dynamically to the router name
                    new_filename = filename.replace("xxx", router_name)
                    old_file = os.path.join(new_folder_path, filename)
                    new_file = os.path.join(new_folder_path, new_filename)

                    if old_file != new_file:
                        os.rename(old_file, new_file)

                    # Only process Excel files containing the SCC keyword
                    if TEMPLATE_FILE_KEYWORD.lower() in new_filename.lower() and new_filename.endswith(".xlsx"):

                        # ✨ THE FIX: Path now matches your exact folder structure! ✨
                        log_dir_path = os.path.join(outputs_base_folder, LOG_SUBFOLDER, site_name)
                        log_file_path = None

                        if os.path.exists(log_dir_path):
                            for potential_file in os.listdir(log_dir_path):
                                if router_name in potential_file:
                                    log_file_path = os.path.join(log_dir_path, potential_file)
                                    break

                        if log_file_path and os.path.exists(log_file_path):
                            full_log_text = safe_read_log(log_file_path)

                            if update_excel_document(new_file, router_name, ip_address, full_log_text):
                                print(f"  -> SUCCESS: Injected mapped commands into Excel cells.")
                            else:
                                print(f"  -> WARNING: Could not find the exact commands written inside the Excel file.")
                        else:
                            print(f"  -> [!] Missing Log File for {router_name} in folder: {LOG_SUBFOLDER}/{site_name}")

            else:
                print(f"\n[-] Skipped: Folder {router_name} already exists. Please delete it to re-run.")

    print("\n✅ Multi-Site SCC Deployment Complete!")


if __name__ == "__main__":
    process_scc_excel()
